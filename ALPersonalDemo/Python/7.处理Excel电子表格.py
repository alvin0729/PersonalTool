

12.2安装openpyxl模块

12.3读取Excel文档
>>> import openpyxl
>>> wb = openpyxl.load_workbook('example.xlsx')
>>> type(wb)
<class 'openpyxl.workbook.workbook.Workbook'>


每个表由一个 Worksheet 对象表示，通过向工作簿方法 get_sheet_by_name()传递表名字符串获得。
import openpyxl
>>> wb = openpyxl.load_workbook('example.xlsx')
>>> wb.get_sheet_names()
>>> sheet = wb.get_sheet_by_name('Sheet1')
>>> sheet
<Worksheet "Sheet1">
>>> type(sheet)
<class 'openpyxl.worksheet.worksheet.Worksheet'>
>>> sheet.title
'Sheet1'
>>> anotherSheet = wb.get_active_sheet()
>>> anotherSheet
<Worksheet "Sheet1">

可以调用 Workbook 对象的 get_active_sheet()方法，取得
工作簿的活动表。活动表是工作簿在 Excel 中打开时出现的工作表。


12.3.3 从表中取得单元格
>>> import openpyxl
>>> wb = openpyxl.load_workbook('example.xlsx')
>>> sheet = wb.get_sheet_by_name('Sheet1')
>>> sheet['A1']
<Cell 'Sheet1'.A1>
>>> sheet['A1'].value
'4/5/2015 1:34:02 PM '
>>> c = sheet['B1']
>>> c.value
'Apples'
>>> 'Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value
'Row 1, Column B is Apples'
>>> 'Cell ' + c.coordinate + ' is ' + c.value
'Cell B1 is Apples'
>>> sheet['C1'].value
73
>>> sheet['C1'].coordinate
'C1'

>>> sheet.cell(row=1, column=2)
<Cell 'Sheet1'.B1>
>>> sheet.cell(row=1, column=2).value
'Apples'
>>> for i in range(1, 8, 2):
...     print(i, sheet.cell(row=i, column=2).value)
... 
1 Apples
3 Pears
5 Apples
7 Strawberries

可以通过 Worksheet 对象的 get_highest_row()和 get_highest_column()方法，确定表的大小。
>>> sheet = wb.get_sheet_by_name('Sheet1')
>>> sheet.max_row
7
>>> sheet.max_column
3

12.4.5 列字母和数字之间的转换
要从字母转换到数字，就调用 openpyxl.cell.column_index_from_string()函数。
要从数字转换到字母，就调用 openpyxl.cell.get_column_letter()函数。


>>> import openpyxl
try: 
    from openpyxl.cell import get_column_letter, column_index_from_string
except ImportError:
    from openpyxl.utils import get_column_letter, column_index_from_string
>>> get_column_letter(1)
'A'
>>> get_column_letter(2)
'B'
>>> get_column_letter(27)
'AA'
>>> get_column_letter(900)
'AHP'
>>> wb = openpyxl.load_workbook('example.xlsx')
>>> sheet = wb.get_sheet_by_name('Sheet1')
>>> get_column_letter(sheet.max_column)
'C'
>>> column_index_from_string('A')
1
>>> column_index_from_string('AA')
27


12.3.5从表中取得行和列
>>> import openpyxl
>>> wb = openpyxl.load_workbook('example.xlsx')
>>> sheet = wb.get_sheet_by_name('Sheet1')
>>> tuple(sheet['A1':'C3'])
((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>), (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>), (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>))
>>> for rowOfCellObjects in sheet['A1':'C3']:
...     for cellObj in rowOfCellObjects:
...          print(cellObj.coordinate, cellObj.value)
...     print('--- END OF ROW ---')


>>> sheet = wb.get_active_sheet()
>>> sheet['B']
(<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.B4>, <Cell 'Sheet1'.B5>, <Cell 'Sheet1'.B6>, <Cell 'Sheet1'.B7>)

>>> for cellObj in sheet['B']:
...     print(cellObj.value)
... 
Apples
Cherries
Pears
Oranges
Apples
Bananas
Strawberries


12.3.6工作簿、工作表、单元格
1．导入 openpyxl 模块。
2．调用 openpyxl.load_workbook()函数。
3．取得 Workbook 对象。
4．调用 get_active_sheet()或 get_sheet_by_name()工作簿方法。
5．取得 Worksheet 对象。
6．使用索引或工作表的 cell()方法，带上 row 和 column 关键字参数。
7．取得 Cell 对象。
8．读取 Cell 对象的 value 属性。


12.4 项目：从电子表格中读取数据

第1步឴：读取电子表格数据

第2步：填充数据结构
保存在 countyData 中的数据结构将是一个字典，以州的简称作为键。每个州的
简称将映射到另一个字典，其中的键是该州的县的名称。每个县的名称又映射到一
个字典，该字典只有两个键， 'tracts'和'pop'。这些键映射到普查区数目和该县的人口。
{'AK': {'Aleutians East': {'pop': 3141, 'tracts': 1},
'Aleutians West': {'pop': 5561, 'tracts': 2},
'Anchorage': {'pop': 291826, 'tracts': 55},
'Bethel': {'pop': 17013, 'tracts': 3},
'Bristol Bay': {'pop': 997, 'tracts': 1},
   --snip--

>>> countyData['AK']['Anchorage']['pop']
291826
>>> countyData['AK']['Anchorage']['tracts']
55

第3步：将结果写入文件
#! python3
# readCensusExcel.py - Tabulates population and number of census tracts for
# each county.

import openpyxl, pprint

print('Opening workbook...')

wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb.get_sheet_by_name('Population by Census Tract')
countyData = {}

# TODO: Fill in countyData with each county's population and tracts.
print('Reading rows...')
for row in range(2, sheet.max_row + 1):
    # Each row in the spreadsheet has data for one census tract.
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    # Make sure the key for this state exists.
    countyData.setdefault(state, {})
    # Make sure the key for this county in this state exists.
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})
    # Each row represents one census tract, so increment by one.
    countyData[state][county]['tracts'] += 1
    # Increase the county pop by the pop in this census tract.
    countyData[state][county]['pop'] += int(pop)

# Open a new text file and write the contents of countyData to it.
print('Writing results...')
resultFile = open('census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()
print('Done.')



12.5写入Excel文档

12.5.1创建并保存Excel文档
调用 openpyxl.Workbook()函数，创建一个新的空 Workbook 对象。

>>> import openpyxl
>>> wb = openpyxl.Workbook()
>>> wb.get_sheet_names()
['Sheet']
>>> sheet = wb.get_active_sheet()
>>> sheet.title
'Sheet'
>>> sheet.title = 'Spam Bacon Eggs Sheet'
>>> wb.get_sheet_names()
['Spam Bacon Eggs Sheet']

>>> wb.save('example_copy.xlsx')
用 save()工作簿方法


12.5.2 创建和删除工作表

create_sheet()方法返回一个新的 Worksheet 对象，名为 SheetX，它默认是工作簿的最后一个工作表。或者，可以利用 index 和 title 关键字参数，指定新工作表的
索引或名称。

>>> import openpyxl
>>> wb = openpyxl.Workbook()
>>> wb.get_sheet_names()
['Sheet']
>>> wb.create_sheet()
<Worksheet "Sheet1">
>>> wb.get_sheet_names()
['Sheet', 'Sheet1']
>>> wb.create_sheet(index=0, title='First Sheet')
<Worksheet "First Sheet">
>>> wb.get_sheet_names()
['First Sheet', 'Sheet', 'Sheet1']
>>> wb.create_sheet(index=2, title='Middle Sheet')
<Worksheet "Middle Sheet">
>>> wb.get_sheet_names()
['First Sheet', 'Sheet', 'Middle Sheet', 'Sheet1']
>>> wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
>>> wb.remove_sheet(wb.get_sheet_by_name('Sheet1'))
>>> wb.get_sheet_names()
['First Sheet', 'Sheet']
>>> 

remove_sheet()方法接受一个 Worksheet 对象作为其参数


12.5.3 将值写入单元格
>>> wb['Sheet']
<Worksheet "Sheet">
>>> sheet['A1'] = 'Hello world!'
>>> sheet['A1'].value
'Hello world!'
>>> wb.save('example1_copy.xlsx')


12.6 项目：更新一个电子表格

第1步：利用更新信息建立数据结构
需要更新的价格如下：
Celery 1.19
Garlic 3.07
Lemon 1.27



第2步：检查所有行，更新不正确的价格

 #! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')


# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
'Celery': 1.19,
'Lemon': 1.27}

# Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row): # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
wb.save('updatedProduceSales.xlsx')


12.7 设置单元格的字体风格

>>> import openpyxl
>>> from openpyxl.styles import Font, NamedStyle
>>> wb = openpyxl.Workbook()
>>> sheet = wb.get_sheet_by_name('Sheet')
>>> italic24Font = Font(size=24, italic=True)
>>> styleObj = NamedStyle(font=italic24Font)
>>> sheet['A'].style/styleObj
>>> sheet['A1'] = 'Hello world!'
>>> wb.save('styled.xlsx')


12.8 Font对象

>>> import openpyxl
>>> from openpyxl.styles import Font, Style
>>> wb = openpyxl.Workbook()
>>> sheet = wb.get_sheet_by_name('Sheet')
>>> fontObj1 = Font(name='Times New Roman', bold=True)
>>> styleObj1 = Style(font=fontObj1)
>>> sheet['A1'].style/styleObj
>>> sheet['A1'] = 'Bold Times New Roman'
>>> fontObj2 = Font(size=24, italic=True)
>>> styleObj2 = Style(font=fontObj2)
>>> sheet['B3'].style/styleObj
>>> sheet['B3'] = '24 pt Italic'
>>> wb.save('styles.xlsx')


12.9 公式

>>> import openpyxl
>>> wb = openpyxl.Workbook()
>>> sheet = wb.get_active_sheet()
>>> sheet['A1'] = 200
>>> sheet['A2'] = 300
>>> sheet['A3'] = '=SUM(A1:A2)'
>>> wb.save('writeFormula.xlsx')


如果你希望看到该公式的计算结果，而不是原来的公式，就必需将 load_workbook()的 data_only 关键字参
数设置为 True。这意味着 Workbook 对象要么显示公式，要么显示公式的结果，不能ެ兼得
>>> import openpyxl
>>> wbFormulas = openpyxl.load_workbook('writeFormula.xlsx')
>>> sheet = wbFormulas.get_active_sheet()
>>> sheet['A3'].value
'=SUM(A1:A2)'
>>> wbDataOnly = openpyxl.load_workbook('writeFormula.xlsx', data_only=True)
>>> sheet = wbDataOnly.get_active_sheet()
>>> sheet['A3'].value
500


12.10 调整行和列

12.10.1 设置行高和列宽

>>> import openpyxl
>>> wb = openpyxl.Workbook()
>>> sheet = wb.get_active_sheet()
>>> sheet['A1'] = 'Tall row'
>>> sheet['B2'] = 'Wide column'
>>> sheet.row_dimensions[1].height = 70
>>> sheet.column_dimensions['B'].width = 20
>>> wb.save('dimensions.xlsx')


12.10.2 合并和拆分单元格

>>> import openpyxl
>>> wb = openpyxl.Workbook()
>>> sheet = wb.get_active_sheet()
>>> sheet.merge_cells('A1:D3')
>>> sheet['A1'] = 'Twelve cells merged together.'
>>> sheet.merge_cells('C5:D5')
>>> sheet['C5'] = 'Two merged cells.'
>>> wb.save('merged.xlsx')

要拆分单元格，就调用 unmerge_cells()工作表方法。
>>> import openpyxl
>>> wb = openpyxl.load_workbook('merged.xlsx')
>>> sheet = wb.get_active_sheet()
>>> sheet.unmerge_cells('A1:D3')
>>> sheet.unmerge_cells('C5:D5')
>>> wb.save('merged.xlsx')


12.10.3 冻结窗格
߫冻结的列或行表头，就算用户滚动电子表格，也是始终可见的。

每个 Worksheet 对象都有一个 freeze_panes属性，可以设置为一个 Cell 对象或一个单元格坐标的字符串。

要解冻所有的单元格，就将 freeze_panes 设置为 None 或'A1'。

freeze_panes 的设置             冻结的行和列
sheet.freeze_panes = 'A2'      行 1
sheet.freeze_panes = 'B1'      列 A
sheet.freeze_panes = 'C1'      列 A 和列 B
sheet.freeze_panes = 'C2'      行 1 和列 A 和列 B
sheet.freeze_panes = 'A1'或    没有߫结窗格
sheet.freeze_panes = None

>>> import openpyxl
>>> wb = openpyxl.load_workbook('produceSales.xlsx')
>>> sheet = wb.get_active_sheet()
>>> sheet.freeze_panes = 'A2'
>>> wb.save('freezeExample.xlsx')


12.10.4 图表

1．从一个矩形区域选择的单元格，创建一个 Reference 对象。
2．通过传入 Reference 对象，创建一个 Series 对象。
3．创建一个 Chart 对象。
4．将 Series 对象添加到 Chart 对象。
5． 可选地设置 Chart 对象的 drawing.top、 drawing.left、 drawing.width 和 drawing.height
变量。
6．将 Chart 对象添加到 Worksheet 对象。

>>> import openpyxl
>>> wb = openpyxl.Workbook()
>>> sheet = wb.get_active_sheet()
>>> for i in range(1, 11): # create some data in column A
sheet['A' + str(i)] = i
>>> refObj = openpyxl.charts.Reference(sheet, (1, 1), (10, 1))
>>> seriesObj = openpyxl.charts.Series(refObj, title='First series')
>>> chartObj = openpyxl.charts.BarChart()
>>> chartObj.append(seriesObj)
>>> chartObj.drawing.top = 50 # set the position
>>> chartObj.drawing.left = 100
>>> chartObj.drawing.width = 300 # set the size
>>> chartObj.drawing.height = 200
>>> sheet.add_chart(chartObj)
>>> wb.save('sampleChart.xlsx')


我们可以调用 openpyxl.charts.BarChart()，创建一个条形图。也可以调用
openpyxl.charts.LineChart()、 openpyxl.charts.ScatterChart()和 openpyxl.charts.PieChart()， 创
建折线图、散点图和饼图。










